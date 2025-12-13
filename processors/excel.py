"""Excel spreadsheet processing utilities."""

import os
from openpyxl import load_workbook
import csv


def process(file_path, use_ocr=None):
    """
    Extract text from Excel file and convert to Markdown.

    Args:
        file_path: Path to .xlsx/.xls/.csv file
        use_ocr: Ignored for Excel (always direct extraction)

    Returns:
        {
            "text": str,
            "file_size": int,
            "num_pages": int,
            "processing_method": str,
            "metadata": dict
        }
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    file_size = os.path.getsize(file_path)
    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.csv':
        return process_csv(file_path, file_size)
    else:
        return process_xlsx(file_path, file_size)


def process_xlsx(file_path, file_size):
    """Process Excel workbook."""
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ValueError(f"Invalid Excel file: {e}")

    md_parts = []
    sheets_data = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        rows = [row for row in rows if any(cell is not None for cell in row)]
        if not rows:
            continue

        max_cols = max(len(row) for row in rows)

        sheet_md = f"## Sheet: {sheet_name}\n\n"

        if len(rows) > 1000:
            sheet_md += f"*Note: Showing first 1000 rows of {len(rows)} total*\n\n"
            display_rows = rows[:1000]
        else:
            display_rows = rows

        header = display_rows[0] if display_rows else []
        header = [str(cell) if cell is not None else '' for cell in header]
        while len(header) < max_cols:
            header.append('')

        sheet_md += '| ' + ' | '.join(header) + ' |\n'
        sheet_md += '|' + '|'.join(['---'] * max_cols) + '|\n'

        for row in display_rows[1:]:
            cells = [str(cell) if cell is not None else '' for cell in row]
            while len(cells) < max_cols:
                cells.append('')
            cells = [c.replace('|', '\\|').replace('\n', ' ') for c in cells]
            sheet_md += '| ' + ' | '.join(cells) + ' |\n'

        md_parts.append(sheet_md)
        sheets_data.append({
            "name": sheet_name,
            "rows": len(rows),
            "columns": max_cols
        })
        total_rows += len(rows)

    wb.close()

    full_text = "# Spreadsheet\n\n" + "\n\n".join(md_parts)

    if not full_text.strip():
        raise ValueError("Could not extract any data from Excel file.")

    return {
        "text": full_text,
        "file_size": file_size,
        "num_pages": len(sheets_data),
        "processing_method": "direct",
        "metadata": {
            "sheets": sheets_data,
            "total_rows": total_rows
        }
    }


def process_csv(file_path, file_size):
    """Process CSV file."""
    rows = []
    encoding = 'utf-8'

    for enc in ['utf-8', 'latin-1', 'cp1252']:
        try:
            with open(file_path, 'r', encoding=enc) as f:
                reader = csv.reader(f)
                rows = list(reader)
                encoding = enc
                break
        except UnicodeDecodeError:
            continue

    if not rows:
        raise ValueError("Could not read CSV file with any encoding.")

    rows = [row for row in rows if any(cell.strip() for cell in row)]
    if not rows:
        raise ValueError("CSV file is empty.")

    max_cols = max(len(row) for row in rows)

    md_parts = ["# CSV Data\n"]

    if len(rows) > 1000:
        md_parts.append(f"*Note: Showing first 1000 rows of {len(rows)} total*\n\n")
        display_rows = rows[:1000]
    else:
        display_rows = rows

    header = display_rows[0] if display_rows else []
    while len(header) < max_cols:
        header.append('')

    md_parts.append('| ' + ' | '.join(header) + ' |')
    md_parts.append('|' + '|'.join(['---'] * max_cols) + '|')

    for row in display_rows[1:]:
        cells = list(row)
        while len(cells) < max_cols:
            cells.append('')
        cells = [c.replace('|', '\\|').replace('\n', ' ') for c in cells]
        md_parts.append('| ' + ' | '.join(cells) + ' |')

    full_text = '\n'.join(md_parts)

    return {
        "text": full_text,
        "file_size": file_size,
        "num_pages": 1,
        "processing_method": "direct",
        "metadata": {
            "rows": len(rows),
            "columns": max_cols,
            "encoding": encoding
        }
    }


def analyze_excel(file_path):
    """
    Analyze Excel file.

    Args:
        file_path: Path to Excel file

    Returns:
        {
            "num_pages": int,
            "ocr_recommended": bool,
            "ocr_reason": str
        }
    """
    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == '.csv':
            return {
                "num_pages": 1,
                "ocr_recommended": False,
                "ocr_reason": "CSV files use direct text extraction"
            }

        wb = load_workbook(file_path, data_only=True)
        num_sheets = len(wb.sheetnames)
        wb.close()

        return {
            "num_pages": num_sheets,
            "ocr_recommended": False,
            "ocr_reason": "Excel files use direct data extraction"
        }
    except Exception as e:
        return {
            "num_pages": 0,
            "ocr_recommended": False,
            "ocr_reason": f"Analysis failed: {str(e)}"
        }
